#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = [
#   "openpyxl>=3.1",
#   "jsonschema>=4.0",
#   "anthropic>=0.40.0,<1.0.0",
#   "youtube-transcript-api>=1.2.4",
#   "yt-dlp>=2026.7.0",
#   "pydantic>=2.0",
# ]
# ///
"""sheet2pack — convert a SODA course-tracker spreadsheet into a LearnForge
importable course pack (exported-course JSON, consumed by Settings → Import).

Input convention (per-course xlsx, e.g. "SFD402 - MLOps Bootcamp.xlsx"):
  - "Tracker" sheet:
      module rows : col A = module number (1.0), col C = module name
      lesson rows : col B = lesson number (101), col C = title,
                    col D = filename, any cell containing a YouTube URL
      resource rows: no lesson number; col C = label, col D = text/URL —
                    appended to the previous lesson's markdown
  - "Quizzes" sheet (optional):
      chapter header rows: col A = "Chapter N - Title"
      question rows: B=S.No, C=Type, D=Question, E-H=Options A-D, I=Answer,
                     J=Explanation
      Questions attach as a module quiz block ONLY when "Chapter N" matches a
      module number from Tracker; unmatched chapters are reported and skipped
      (guards against stale template sheets reused across course workbooks).

Output: exported-course JSON (exportVersion present → relaxed schema branch),
videos keyed by SECTION id (= block id) so per-lesson hero videos display
after import (requires the section-keyed video fix, commit 98c3fcf).

Usage:
  python3 scripts/sheet2pack.py "<course>.xlsx" -o out.json \
      [--id sfd402-mlops] [--title "MLOps Bootcamp"] [--domain devops] \
      [--channel "School of Devops"] [--enrich] [--yes]

Enrichment (--enrich):
  Fetches YouTube transcripts and calls the Anthropic API to generate
  transcript-grounded lesson text and MCQ quizzes. Requires ANTHROPIC_API_KEY.

Validation: if `jsonschema` is installed, the output is validated against
learnforge-core/topic-packs/pack-schema.json before writing.
"""

import argparse
import asyncio
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

YT_RE = re.compile(r"(?:youtu\.be/|youtube\.com/watch\?v=)([A-Za-z0-9_-]{6,20})")
NOW = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

# R1 caps from pack-schema.json
MAX_OBJECTIVES = 12
MAX_OBJECTIVE_LEN = 280
MAX_TITLE_LEN = 200
MAX_MODULE_DESC_LEN = 1000


def cell(row, i):
    v = row[i] if i < len(row) else None
    return str(v).strip() if v is not None else ""


def find_yt(row):
    for v in row:
        if v is None:
            continue
        m = YT_RE.search(str(v))
        if m:
            return m.group(1)
    return None


def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        return False


def parse_tracker(ws):
    """Return ([{num, title, lessons: [...]}], warnings)."""
    modules = []
    warnings = []
    current = None
    last_lesson = None
    for row in ws.iter_rows(values_only=True):
        if not row or not any(v is not None for v in row):
            continue
        a, b, c, d = cell(row, 0), cell(row, 1), cell(row, 2), cell(row, 3)
        # module row: numeric col A (name may be missing — placeholder + warn,
        # otherwise its lessons would silently leak into the previous module)
        if a and is_number(a) and not b:
            num = int(float(a))
            title = c or f"Module {num}"
            if not c:
                warnings.append(f"module {num} has no name in Tracker — using placeholder \"{title}\"")
            current = {"num": num, "title": title, "lessons": []}
            modules.append(current)
            last_lesson = None
            continue
        if current is None:
            continue  # preamble/stats rows before first module
        # lesson row: numeric col B + a title in col C
        if b and is_number(b) and c:
            last_lesson = {
                "num": int(float(b)),
                "title": c,
                "filename": d,
                "video_id": find_yt(row),
                "resources": [],
            }
            current["lessons"].append(last_lesson)
            continue
        # resource row: no lesson number, but has label text — attach to prev lesson
        if not b and c and last_lesson is not None:
            last_lesson["resources"].append({"label": c, "text": d})
    return [m for m in modules if m["lessons"]], warnings


def title_overlap(a, b):
    """Token overlap ratio between two titles (0..1), stopwords ignored."""
    stop = {"the", "a", "an", "and", "or", "to", "of", "with", "for", "in", "on"}
    ta = {w for w in re.findall(r"[a-z]+", a.lower()) if w not in stop}
    tb = {w for w in re.findall(r"[a-z]+", b.lower()) if w not in stop}
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / min(len(ta), len(tb))


def parse_quizzes(ws):
    """Return {chapter_num: {"title": str, "questions": [...]}}."""
    chapters = {}
    current_num = None
    chap_re = re.compile(r"Chapter\s+(\d+)\s*-?\s*(.*)", re.I)
    for row in ws.iter_rows(values_only=True):
        if not row or not any(v is not None for v in row):
            continue
        a = cell(row, 0)
        m = chap_re.search(a) if a else None
        if m:
            current_num = int(m.group(1))
            chapters.setdefault(current_num, {"title": m.group(2).strip(), "questions": []})
            continue
        stem = cell(row, 3)
        if current_num is None or not stem or stem == "Question":
            continue
        options = [cell(row, i) for i in (4, 5, 6, 7)]
        options = [o for o in options if o]
        answer = cell(row, 8)
        if not options or not answer:
            continue
        chapters[current_num]["questions"].append(
            {"stem": stem, "options": options, "answer": answer,
             "explanation": cell(row, 9)}
        )
    return {k: v for k, v in chapters.items() if v["questions"]}


def lesson_markdown(lesson):
    # No H1 — the app already renders the lesson title from params.lesson_title.
    parts = []
    if lesson["video_id"]:
        parts.append(
            "▶ **Video lesson** — the reference video for this lesson plays in the panel above."
        )
    for r in lesson["resources"]:
        if r["text"]:
            if r["text"].startswith("http"):
                parts.append(f"**{r['label']}**: <{r['text']}>")
            else:
                parts.append(f"**{r['label']}** — {r['text']}")
        else:
            parts.append(f"**{r['label']}**")
    if not parts:
        parts.append("_Video for this lesson is coming soon._")
    return "\n\n".join(parts)


def quiz_payload(questions, module_slug):
    qs = []
    for qi, q in enumerate(questions, 1):
        opts = [{"id": f"opt-{qi}-{oi}", "text": o} for oi, o in enumerate(q["options"], 1)]
        correct = next(
            (o["id"] for o in opts if o["text"].strip().lower() == q["answer"].strip().lower()),
            None,
        )
        if correct is None:
            continue  # answer text doesn't match any option — skip defensively
        qs.append(
            {
                "id": f"q-{module_slug}-{qi}",
                "stem": q["stem"],
                "options": opts,
                "correctOptionId": correct,
                "explanation": q.get("explanation") or "",
            }
        )
    return {"questions": qs} if qs else None


def block(bid, module_id, ordering, block_type, payload, params=None):
    return {
        "id": bid,
        "moduleId": module_id,
        "ordering": ordering,
        "blockType": block_type,
        "status": "ready",
        "paramsJson": json.dumps(params or {}, ensure_ascii=False),
        "payloadJson": json.dumps(payload, ensure_ascii=False),
        "sourceAnchorsJson": "[]",
        "metadataJson": "{}",
        "retryCount": 0,
        "createdAt": NOW,
        "updatedAt": NOW,
    }


def compute_matched_chapters(modules_raw, quizzes):
    """Return the set of module nums whose hand-authored quiz title overlaps >= 0.5.

    Extracted from the assembly loop so the D-09 fill-gaps decision is computed
    once and reused in both the stale-quiz guard (assembly) and the enrichment
    pre-pass (quiz generation).
    """
    matched = set()
    for m in modules_raw:
        if m["num"] in quizzes:
            ch = quizzes[m["num"]]
            if title_overlap(ch["title"], m["title"]) >= 0.5:
                slug = f"mod-{m['num']}"
                qp = quiz_payload(ch["questions"], slug)
                if qp:
                    matched.add(m["num"])
    return matched


def assemble_payload(
    modules_raw,
    quizzes,
    enriched_lessons,
    generated_quizzes,
    channel,
    pack_id,
    title,
    domain,
    licensed=False,
    licensor=None,
    return_warnings=False,
):
    """Assemble the exported-course JSON payload from parsed modules and enrichment data.

    This function is the single implementation of the assembly loop that
    convert() calls after the enrichment pre-pass. It is also used directly
    by tests so they can inject enriched_lessons/generated_quizzes without
    hitting the network (ENR-04 / D-03 / D-07 / D-09 / D-16 integration tests).

    Args:
        modules_raw:        List of module dicts from parse_tracker().
        quizzes:            Dict of hand-authored quiz chapters from parse_quizzes().
        enriched_lessons:   {video_id: markdown} from generate_lessons() or {}.
        generated_quizzes:  {module_slug: quiz_payload_dict} from generate_quizzes() or {}.
        channel:            YouTube channel name for video metadata.
        pack_id:            Unique pack identifier slug.
        title:              Human-readable course title.
        domain:             Domain/category tag for the pack.
        licensed:           If True, stamp exportedFrom as "licensed:{pack_id}|{licensor}"
                            (paid pack, non-exportable after import). Default False stamps
                            "imported:{pack_id}".
        licensor:           Display name of the licensing entity. When licensed=True and
                            licensor is None or empty, defaults to channel. Any "|" in the
                            effective licensor is replaced with "/" to keep the pipe-split
                            unambiguous (T-g73-02). Ignored when licensed=False.
        return_warnings:    If True, return (payload, warnings) tuple instead of payload.

    Returns:
        payload dict, or (payload, warnings) tuple if return_warnings=True.
    """
    modules, blocks_map, videos_map = [], {}, {}
    warnings: list = []
    # Track which chapters got hand-authored quizzes (for D-09 fill-gaps + report)
    used_chapters: set = set()

    for m in modules_raw:
        slug = f"mod-{m['num']}"
        lesson_titles = [l["title"] for l in m["lessons"]]
        objectives = [t[:MAX_OBJECTIVE_LEN] for t in lesson_titles[:MAX_OBJECTIVES]]
        desc = f"{len(m['lessons'])} video lessons: " + "; ".join(lesson_titles[:4])
        modules.append(
            {
                "id": slug,
                "title": m["title"][:MAX_TITLE_LEN],
                "description": desc[:MAX_MODULE_DESC_LEN],
                "objectives": objectives,
            }
        )

        mblocks = []
        for i, lesson in enumerate(m["lessons"]):
            bid = f"blk-{lesson['num']}"
            vid = lesson.get("video_id")

            # D-07: video always stays in videos_map regardless of enrichment
            if vid:
                videos_map[bid] = [
                    {
                        "videoId": vid,
                        "title": lesson["title"][:500],
                        "channelTitle": channel,
                        "relevanceScore": 1.0,
                    }
                ]
            else:
                warnings.append(f"no video: {m['num']}/{lesson['num']} {lesson['title']}")

            # Select markdown source: enriched (if generated) or stage-1 stub (D-03)
            if vid and vid in enriched_lessons:
                md = enriched_lessons[vid]
                # Route through the 131072-byte guard — single implementation in
                # lesson_generator._make_section_payload (no duplicate guard here)
                from enrichment.lesson_generator import _make_section_payload
                payload_json_str = _make_section_payload(md)
                section_payload = json.loads(payload_json_str)
            else:
                # D-03: skipped/caption-less lesson keeps the stage-1 video-only stub
                section_payload = {"markdown": lesson_markdown(lesson)}

            mblocks.append(
                block(
                    bid, slug, i, "section",
                    section_payload,
                    # LessonNavList reads params.lesson_title for the sidebar label
                    params={"lesson_title": lesson["title"][:MAX_TITLE_LEN]},
                )
            )

        # Attach a chapter quiz ONLY when both the number AND the title agree —
        # quiz sheets are often stale template leftovers from other courses.
        if m["num"] in quizzes:
            ch = quizzes[m["num"]]
            if title_overlap(ch["title"], m["title"]) >= 0.5:
                qp = quiz_payload(ch["questions"], slug)
                if qp:
                    mblocks.append(block(f"quiz-{slug}", slug, len(mblocks), "quiz", qp))
                    used_chapters.add(m["num"])
            else:
                warnings.append(
                    f"quiz chapter {m['num']} \"{ch['title'][:40]}\" != module \"{m['title'][:40]}\" — skipped (stale template sheet?)"
                )

        # D-09: fill gaps only — add generated quiz if no hand-authored one matched
        if slug in generated_quizzes and m["num"] not in used_chapters:
            mblocks.append(
                block(f"quiz-{slug}", slug, len(mblocks), "quiz", generated_quizzes[slug])
            )

        blocks_map[slug] = mblocks

    for ch_num in sorted(set(quizzes) - used_chapters):
        if not any(m["num"] == ch_num for m in modules_raw):
            warnings.append(f"quiz chapter {ch_num} matches no module number — skipped")

    edges = [
        {"from": modules[i]["id"], "to": modules[i + 1]["id"]}
        for i in range(len(modules) - 1)
    ]

    payload = {
        "id": pack_id,
        "title": title,
        "description": f"{title} — converted from course tracker ({len(modules)} modules, "
        f"{sum(len(m['lessons']) for m in modules_raw)} lessons).",
        "domain_module": domain,
        "modules": modules,
        "edges": edges,
        "exportVersion": "1.0.0",
        "exportedAt": NOW,
        "exportedFrom": (
            f"licensed:{pack_id}|{(licensor or channel).replace('|', '/')}"
            if licensed
            else f"imported:{pack_id}"
        ),
        "blocks": blocks_map,
        "videos": videos_map,
    }

    if return_warnings:
        return payload, warnings
    return payload


def convert(xlsx_path, pack_id, title, domain, channel,
            enrich=False, enrich_only=None, yes=False, licensed=False, licensor=None):
    """Convert a SODA xlsx to an exported-course JSON pack.

    Args:
        xlsx_path:    Path to the course tracker xlsx.
        pack_id:      Unique pack identifier slug.
        title:        Human-readable course title.
        domain:       Domain/category tag for the pack.
        channel:      YouTube channel name for video metadata.
        enrich:       Run all enrichment stages (transcripts, lessons, quizzes).
        enrich_only:  Run only one stage: 'transcripts', 'lessons', or 'quizzes'.
        yes:          Skip the D-15 cost-confirmation prompt.
        licensed:     If True, stamp exportedFrom as "licensed:{pack_id}|{licensor}" (paid pack).
        licensor:     Display name of licensing entity; defaults to channel when licensed=True.

    Returns:
        (payload_dict, warnings, skipped, failed) — the last two lists are
        populated only when enrich or enrich_only is set.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Tracker" not in wb.sheetnames:
        sys.exit("ERROR: no 'Tracker' sheet found")
    modules_raw, parse_warnings = parse_tracker(wb["Tracker"])
    if not modules_raw:
        sys.exit("ERROR: no modules with lessons found in Tracker")
    quizzes = parse_quizzes(wb["Quizzes"]) if "Quizzes" in wb.sheetnames else {}

    # Enrichment state — populated in the pre-pass below
    enriched_lessons: dict = {}   # video_id → generated markdown text
    generated_quizzes: dict = {}  # module_slug → quiz_payload dict
    skipped: list = []            # lessons with no captions (D-03)
    failed: list = []             # lessons/modules that failed generation (D-16)

    # Pre-compute matched chapters for D-09 fill-gaps guard
    matched_chapters: set = compute_matched_chapters(modules_raw, quizzes)

    # ------------------------------------------------------------------
    # Enrichment pre-pass (runs before assembly; no network calls otherwise)
    # ------------------------------------------------------------------
    if enrich or enrich_only:
        from enrichment.transcript import fetch_and_cache_transcripts
        from enrichment.token_estimator import estimate_and_confirm
        from enrichment.quiz_generator import generate_quizzes

        # Stage A: fetch transcripts for all lessons with a video_id (D-01 / D-04)
        # Always run — transcripts feed both lessons and quizzes.
        transcripts = fetch_and_cache_transcripts(modules_raw, skipped)

        # Determine which LLM stages to run
        run_lessons = enrich or enrich_only == "lessons"
        run_quizzes = enrich or enrich_only == "quizzes"

        if (run_lessons or run_quizzes) and transcripts:
            # D-15: print estimate and confirm before any LLM calls
            estimate_and_confirm(transcripts, len(modules_raw), yes=yes)

        async def _run_enrichment():
            """Single coroutine wrapping all async enrichment stages (Pitfall 7)."""
            el: dict = {}
            gq: dict = {}
            if run_lessons and transcripts:
                from enrichment.lesson_generator import generate_lessons
                el = await generate_lessons(transcripts, failed)
            if run_quizzes:
                gq = await generate_quizzes(
                    modules_raw, transcripts, matched_chapters, failed
                )
            return el, gq

        # Single entry point for all async enrichment (Pitfall 7 — never nest event loops)
        enriched_lessons, generated_quizzes = asyncio.run(_run_enrichment())

    # ------------------------------------------------------------------
    # Assembly — delegate to assemble_payload() (shared with tests)
    # ------------------------------------------------------------------
    payload, assembly_warnings = assemble_payload(
        modules_raw=modules_raw,
        quizzes=quizzes,
        enriched_lessons=enriched_lessons,
        generated_quizzes=generated_quizzes,
        channel=channel,
        pack_id=pack_id,
        title=title,
        domain=domain,
        licensed=licensed,
        licensor=licensor,
        return_warnings=True,
    )
    warnings = parse_warnings + assembly_warnings
    return payload, warnings, skipped, failed


def validate(payload, schema_path):
    try:
        import jsonschema
    except ImportError:
        return "jsonschema not installed — skipped (pip install jsonschema)"
    schema = json.loads(Path(schema_path).read_text())
    jsonschema.validate(payload, schema)
    return "schema valid (draft 2020-12)"


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("xlsx")
    ap.add_argument("-o", "--out", required=True)
    ap.add_argument("--id", dest="pack_id", default=None)
    ap.add_argument("--title", default=None)
    ap.add_argument("--domain", default="devops")
    ap.add_argument("--channel", default="School of Devops")
    ap.add_argument("--enrich", action="store_true",
                    help="fetch transcripts + generate lessons + quizzes via Anthropic API")
    ap.add_argument("--enrich-only", dest="enrich_only",
                    choices=["transcripts", "lessons", "quizzes"],
                    metavar="STAGE",
                    help="rerun a single enrichment stage only (transcripts, lessons, or quizzes)")
    ap.add_argument("--yes", action="store_true",
                    help="skip cost-confirmation prompt (for scripted runs, D-15)")
    ap.add_argument("--licensed", action="store_true",
                    help="stamp exportedFrom as licensed:{id}|{licensor} — paid pack, non-exportable after import")
    ap.add_argument("--licensor", metavar="NAME", default=None,
                    help="display name of the licensing entity; defaults to --channel when --licensed is set")
    args = ap.parse_args()

    stem = Path(args.xlsx).stem
    title = args.title or re.sub(r"^[A-Z]{2,4}\d+\s*-\s*", "", stem)
    pack_id = args.pack_id or re.sub(r"[^a-z0-9-]+", "-", title.lower()).strip("-")

    payload, warnings, skipped, failed = convert(
        args.xlsx, pack_id, title, args.domain, args.channel,
        enrich=args.enrich,
        enrich_only=args.enrich_only,
        yes=args.yes,
        licensed=args.licensed,
        licensor=args.licensor,
    )

    schema_path = Path(__file__).resolve().parent.parent / "learnforge-core/topic-packs/pack-schema.json"
    vmsg = validate(payload, schema_path) if schema_path.exists() else "schema file not found — skipped"

    # Write the pack BEFORE printing report so a write failure never blocks it (D-16)
    Path(args.out).write_text(json.dumps(payload, ensure_ascii=False, indent=2))

    n_lessons = sum(len(b) for b in payload["blocks"].values())
    n_videos = sum(len(v) for v in payload["videos"].values())
    print(f"pack: {pack_id} — \"{title}\"")
    print(f"modules: {len(payload['modules'])}  blocks: {n_lessons}  videos: {n_videos}")
    print(f"validation: {vmsg}")
    print(f"wrote: {args.out}")
    if warnings:
        print(f"\nwarnings ({len(warnings)}):")
        for w in warnings:
            print(f"  - {w}")
    # D-16 enrichment report
    if skipped:
        print(f"\nskipped ({len(skipped)}) — no captions, stays video-only:")
        for s in skipped:
            print(f"  - {s}")
    if failed:
        print(f"\nfailed ({len(failed)}) — kept as video-only stub:")
        for f_item in failed:
            print(f"  - {f_item}")


if __name__ == "__main__":
    main()
